from openpyxl import Workbook

wb = Workbook()
# wb.save(filename="wb_test2.xlsx")
ws = wb.active
print(ws)

# creating some worksheet

ws1 = wb.create_sheet("sheet1")
ws2 = wb.create_sheet("sheet2")
ws3 = wb.create_sheet("sheet3")
#
# wb.save(filename="wb_test2.xlsx")
# print(wb.worksheets)
#
# ws3.title = "sheet3_changed"
# print(wb.worksheets)
#
# ws3.sheet_properties.tabColor = "1072BA"
#
# wb.save(filename="wb_test2.xlsx")
# ws3 = wb["sheet3_changed"]
# wb.save(filename="wb_test2.xlsx")
#
# print(wb.worksheets)
# print(ws3)
#
# # looping in Workbook
#
# for sheet in wb:
#     print(sheet.title)
#
# # copying worksheet in workbook using Workbook.copy_worksheet() classMethod
#
# ws5 = wb.active
# print(ws5)
# ws6 = wb.copy_worksheet(ws5)
# print(wb.worksheets)
#
# wb.copy_worksheet(ws1)
# print(wb.worksheets)
#
# accessing a cell in woekseet and imput some data date
# c = ws["C4"]
#
# ws["C4"] = 5
# print(c)
# print(c.value)
#
# wb.save("wb_test2.xlsx")
#
# d = ws.cell(row=2, column=2, value=50)
# wb.save("wb_test2.xlsx")
# # print(d)
#
# # for looping while accessing some cells
#
# # for cell in range(1, 7):
# #     print(ws.cell(row=3, column=5, value=90))
#
# wb.save(filename="wb_test2.xlsx")
#
# # accessing one column and one range in worksheet
#
# colB = ws["B"]
#
# # for c in range(1, 10):
# #     print(colB)
# # wb.save(filename="wb_test2.xlsx")
#
# for looping range_cell in worksheet using worksheet.iter_rows() method

# for row in ws.iter_rows(min_row=1, max_col=2, max_row=10):
#     print("test2", row)
#     for cell in row:
#         print("test", cell)
# # looping all rows or columns in file
#
# tuple(wb.rows):

# print(wb.worksheets)
# for x in ws.iter_rows(min_row=1, max_col=1, max_row=10):
#     print(x)
#     c = ws["A1"]
#     ws["A1"] = 10
#     ws.append(x)
#     # print(rows.value)
#     print(c.value)
#     # for value in rows:
#     #     ws["rows"] = 10
# wb.save(filename="wb_test2.xlsx")

data =[(1, "akash deep", "singh"), (2, "avinash kumar", "singh"), (3, " anurag kumar", "singh")]

for cell in data:
    print(cell)
    ws.append(cell)

wb.save(filename="test4.xlsx")