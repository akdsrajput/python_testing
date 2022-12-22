import openpyxl

my_wb = openpyxl.load_workbook("inventory.xlsx")
my_ws = my_wb["Sheet1"]
# print(my_ws)
key = "akash"
# supplier = "AAA Company"

# creating Dictionary and working with  it
total_each_supplier = {}
product_with_less_than_10 = {}
company_with_its_value = {}
# total_each_supplier["key1"] = "A"
# total_each_supplier["key2"] = "B"
# total_each_supplier[key] = "ksjfklsdj"
print(type(total_each_supplier))
# print(total_each_supplier)
# print(total_each_supplier.get(key))
# if supplier in total_each_supplier:
#     print("yes", total_each_supplier.get(supplier))

print(my_ws.max_row)
print(my_ws["D2"].value)
count = 0
for row in range(2, my_ws.max_row + 1):
    print(row, my_ws.cell(row, 4).value)
    supplier = my_ws.cell(row, 4).value
    inventory = my_ws.cell(row, 2).value
    product_number = my_ws.cell(row, 1).value

    # Calculating total number of supplier in my xl file
    if supplier in total_each_supplier:
        total_each_supplier[supplier] = total_each_supplier.get(supplier) + 1

    else:
        total_each_supplier[supplier] = 1

    # Calculating Inventory less than 10
    if inventory < 10:
        product_with_less_than_10[product_number] = inventory
    # else:
    #     continue

    # list each company with respective total value
    price = my_ws.cell(row, 3).value
    if supplier in company_with_its_value:
        company_with_its_value[supplier] = company_with_its_value.get(supplier) + inventory * price
    else:
        company_with_its_value[supplier] = inventory * price

    # write to the next column the value of inventory with respective to its supplier
    my_ws.cell(row, 5).value = inventory * price


print(total_each_supplier)
print(product_with_less_than_10)
print(company_with_its_value)
my_ws.cell(1, 5).value = "value of Inventory"
my_wb.save(filename="inventory.xlsx")