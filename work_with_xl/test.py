from openpyxl import Workbook, load_workbook


wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("sheet1")
wb.save(filename="akash.xlsx")
print(ws)

# wb2 = Workbook()
# ws2 = wb2.create_sheet("akash")
# print(ws2)
# wb2.save("workbook2")
#
# wb3 = load_workbook("inventory.xlsx")
# print(wb3.sheetnames)

wb.create_sheet("wb1_sheet2")
# wb.save(filename="akash.xlsx")
ws3 = wb.create_sheet("wb1_sheet3")
wb.save(filename="akash.xlsx")
print(wb.worksheets)
ws1 = wb.active
print(ws1)

ws.title = "wb1_sheet4"
wb.save(filename="akash.xlsx")
print(wb.worksheets)
# ws4 = wb.wb1_sheet3
print(ws3)
ws3.sheet_properties.tabcolor ="1072BA"